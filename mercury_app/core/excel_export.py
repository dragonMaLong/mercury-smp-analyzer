from __future__ import annotations

import re
from pathlib import Path
from typing import Sequence

from .metrics import summary_metrics
from .models import MercuryResult


def export_results_xlsx(results: Sequence[MercuryResult], output_path: str | Path) -> Path:
    """将选中的计算结果导出到一个 Excel 工作簿。"""
    if not results:
        raise ValueError("没有选中的 SMP 结果可导出。")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("导出 XLS 需要 openpyxl。请运行：python -m pip install -r requirements.txt") from exc

    output = Path(output_path)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"

    header_fill = PatternFill("solid", fgColor="E0ECFF")
    title_font = Font(bold=True, size=12, color="111827")
    header_font = Font(bold=True, color="111827")
    center_alignment = Alignment(horizontal="center")

    _write_summary_sheet(summary, results, header_fill, header_font, title_font, center_alignment)

    used_names = {"汇总"}
    for result in results:
        sheet = workbook.create_sheet(_safe_sheet_name(result.metadata.get("file_name") or result.sample_name, used_names))
        _write_result_sheet(sheet, result, header_fill, header_font, title_font, center_alignment)

    for sheet in workbook.worksheets:
        _autofit_columns(sheet, get_column_letter)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def _write_summary_sheet(sheet, results, header_fill, header_font, title_font, center_alignment) -> None:
    sheet["A1"] = "压汞数据导出"
    sheet["A1"].font = title_font
    sheet["A3"] = "说明"
    sheet["B3"] = "仅导出蓝色圆点选中的样品。dV/dlogD 使用平滑后的入汞孔径分布。"

    headers = [
        "文件",
        "样品名称",
        "测试人员",
        "提交者",
        "创建时间",
        "修改时间",
        "测试仪器",
        "测试软件",
        "进汞接触角 (deg)",
        "表面张力 (dynes/cm)",
        "汞温度 (C)",
        "汞密度 (g/mL)",
        "总入汞体积 (mL/g)",
        "总孔面积 (m2/g)",
        "中值孔径（体积）(nm)",
        "中值压力（体积）(psia)",
        "中值孔径（面积）(nm)",
        "中值压力（面积）(psia)",
        "平均孔径 4V/A (nm)",
        "0.50 psia 体积密度 (g/mL)",
        "表观骨架密度 (g/mL)",
        "孔隙率 (%)",
        "最大压力 (psia)",
        "数据点数",
    ]
    start_row = 5
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for row, result in enumerate(results, start=start_row + 1):
        summary_metrics_result = summary_metrics(result)
        sheet.cell(row, 1, _text(result.metadata.get("file_name")))
        sheet.cell(row, 2, _text(result.metadata.get("sample_name")))
        sheet.cell(row, 3, _text(result.metadata.get("operator")))
        sheet.cell(row, 4, _text(result.metadata.get("submitter")))
        sheet.cell(row, 5, _text(result.metadata.get("created")))
        sheet.cell(row, 6, _text(result.metadata.get("modified")))
        sheet.cell(row, 7, _text(result.metadata.get("instrument_name")))
        sheet.cell(row, 8, _software_text(result))
        sheet.cell(row, 9, _number(result.metadata.get("adv_contact_angle_deg")))
        sheet.cell(row, 10, _number(result.metadata.get("surface_tension_dynes_cm")))
        sheet.cell(row, 11, _number(result.metadata.get("mercury_temperature_C")))
        sheet.cell(row, 12, _number(result.metadata.get("mercury_density_gmL")))
        sheet.cell(row, 13, summary_metrics_result.total_intrusion_volume)
        sheet.cell(row, 14, summary_metrics_result.total_pore_area)
        sheet.cell(row, 15, summary_metrics_result.median_volume_diameter)
        sheet.cell(row, 16, summary_metrics_result.median_volume_pressure)
        sheet.cell(row, 17, summary_metrics_result.median_area_diameter)
        sheet.cell(row, 18, summary_metrics_result.median_area_pressure)
        sheet.cell(row, 19, summary_metrics_result.average_pore_diameter)
        sheet.cell(row, 20, summary_metrics_result.bulk_density)
        sheet.cell(row, 21, summary_metrics_result.apparent_density)
        sheet.cell(row, 22, summary_metrics_result.porosity)
        sheet.cell(row, 23, result.max_pressure)
        sheet.cell(row, 24, result.data_point_count)

    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A{start_row}:X{start_row + len(results)}"


def _write_result_sheet(sheet, result: MercuryResult, header_fill, header_font, title_font, center_alignment) -> None:
    sheet["A1"] = _text(result.metadata.get("file_name"))
    sheet["A1"].font = title_font
    summary_metrics_result = summary_metrics(result)

    metadata_rows = [
        ("样品名称", _text(result.metadata.get("sample_name"))),
        ("测试人员", _text(result.metadata.get("operator"))),
        ("提交者", _text(result.metadata.get("submitter"))),
        ("创建时间", _text(result.metadata.get("created"))),
        ("修改时间", _text(result.metadata.get("modified"))),
        ("测试仪器", _text(result.metadata.get("instrument_name"))),
        ("测试软件", _software_text(result)),
        ("进汞接触角 (deg)", _number(result.metadata.get("adv_contact_angle_deg"))),
        ("表面张力 (dynes/cm)", _number(result.metadata.get("surface_tension_dynes_cm"))),
        ("汞温度 (C)", _number(result.metadata.get("mercury_temperature_C"))),
        ("汞密度 (g/mL)", _number(result.metadata.get("mercury_density_gmL"))),
        ("总入汞体积 (mL/g)", summary_metrics_result.total_intrusion_volume),
        ("总孔面积 (m2/g)", summary_metrics_result.total_pore_area),
        ("中值孔径（体积）(nm)", summary_metrics_result.median_volume_diameter),
        ("中值压力（体积）(psia)", summary_metrics_result.median_volume_pressure),
        ("中值孔径（面积）(nm)", summary_metrics_result.median_area_diameter),
        ("中值压力（面积）(psia)", summary_metrics_result.median_area_pressure),
        ("平均孔径 4V/A (nm)", summary_metrics_result.average_pore_diameter),
        ("0.50 psia 体积密度 (g/mL)", summary_metrics_result.bulk_density),
        ("表观骨架密度 (g/mL)", summary_metrics_result.apparent_density),
        ("孔隙率 (%)", summary_metrics_result.porosity),
        ("最大压力 (psia)", result.max_pressure),
    ]
    for row, (name, value) in enumerate(metadata_rows, start=3):
        sheet.cell(row, 1, name).font = header_font
        sheet.cell(row, 2, value)

    table_row = len(metadata_rows) + 5
    headers = [
        "点号",
        "过程",
        "压力 (psia)",
        "孔径 (nm)",
        "累计孔体积 (mL/g)",
        "增量孔体积 (mL/g)",
        "dV/dlogD（平滑，仅入汞）",
        "总入汞体积占比 (%)",
        "增量入汞体积占比 (%)",
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(table_row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for row_index, row_data in enumerate(result.table, start=table_row + 1):
        is_extrusion = row_data["is_extrusion"] >= 0.5
        sheet.cell(row_index, 1, row_index - table_row)
        sheet.cell(row_index, 2, "退汞" if is_extrusion else "入汞")
        sheet.cell(row_index, 3, row_data["pressure"])
        sheet.cell(row_index, 4, row_data["diameter"])
        sheet.cell(row_index, 5, row_data["cum_volume"])
        sheet.cell(row_index, 6, row_data["incremental_volume"])
        sheet.cell(row_index, 7, None if is_extrusion else max(0.0, row_data["log_diff_intrusion_smooth"]))
        sheet.cell(row_index, 8, row_data["pct_total"])
        sheet.cell(row_index, 9, row_data["pct_incremental"])

    last_row = table_row + len(result.table)
    sheet.freeze_panes = f"A{table_row + 1}"
    sheet.auto_filter.ref = f"A{table_row}:I{last_row}"


def _safe_sheet_name(value, used_names: set[str]) -> str:
    base = _text(value) or "样品"
    base = re.sub(r"[\[\]:*?/\\]", "_", base).strip("'") or "样品"
    base = base[:31]
    name = base
    counter = 2
    while name in used_names:
        suffix = f"_{counter}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(name)
    return name


def _autofit_columns(sheet, get_column_letter) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)


def _text(value) -> str:
    text = str(value or "").strip()
    return text if text else "未记录"


def _software_text(result: MercuryResult) -> str:
    name = str(result.metadata.get("analysis_software") or "").strip()
    version = str(result.metadata.get("analysis_software_version") or "").strip()
    if name and version:
        return f"{name} {version}"
    if name:
        return name
    return _text(result.metadata.get("software_version"))


def _number(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
