from __future__ import annotations

from pathlib import Path
import sys

import numpy as np
import pytest


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = PROJECT_ROOT if (PROJECT_ROOT / "M5G2.SMP").exists() else PROJECT_ROOT.parent
sys.path.insert(0, str(PROJECT_ROOT))

from mercury_app.core import calculate_microactive, export_results_xlsx, load_smp, metrics_for_pressure_range
from mercury_app.core.smp_parser import parse_smp


pytestmark = pytest.mark.skipif(
    not (DATA_ROOT / "M5G2.SMP").exists(),
    reason="Local SMP reference files are not included in the source repository.",
)


COMMON_SMP_FILES = [
    "M5G2.SMP",
    "1-6.SMP",
    "2-M.SMP",
    "BASF400.SMP",
    "JP400.SMP",
    "SEC26.SMP",
    "UMC-50PtA.SMP",
    "PD25002.SMP",
    "Y0509.SMP",
    "Z-Z-1.SMP",
]


def test_load_smp_returns_stable_core_api() -> None:
    result = load_smp(DATA_ROOT / "M5G2.SMP")

    assert result.metadata["file_name"] == "M5G2"
    assert result.metadata["sample_name"] == ""
    assert result.metadata["instrument_name"] == "AutoPore V9600"
    assert result.metadata["analysis_software"] == "MicroActive"
    assert result.metadata["analysis_software_version"] == "2.03.00"
    assert result.data_point_count == len(result.table)
    assert result.pressure.shape == result.diameter.shape == result.cum_volume.shape
    assert result.total_pore_volume > 0
    assert result.max_pressure > 0
    assert np.all(np.isfinite(result.pressure))
    assert np.all(result.diameter > 0)


@pytest.mark.parametrize("filename", COMMON_SMP_FILES)
def test_common_smp_files_do_not_crash(filename: str) -> None:
    result = load_smp(DATA_ROOT / filename)

    assert result.data_point_count > 50
    assert result.total_pore_volume > 0
    assert result.max_pressure > 1000


def test_pressure_range_metrics_update_from_core_only() -> None:
    result = load_smp(DATA_ROOT / "M5G2.SMP")
    metrics = metrics_for_pressure_range(result, 1000, 10000)

    assert metrics.point_count > 0
    assert metrics.pore_volume >= 0
    assert metrics.pore_volume_percent >= 0
    assert metrics.diameter_min <= metrics.diameter_max


def test_contact_angle_override_recalculates_diameter_only() -> None:
    smp = parse_smp(DATA_ROOT / "M5G2.SMP")
    baseline = calculate_microactive(smp)
    edited = calculate_microactive(smp, adv_contact_angle_deg=140.0)

    assert np.allclose(baseline.pressure, edited.pressure)
    assert np.allclose(baseline.cum_volume, edited.cum_volume)
    assert not np.allclose(baseline.diameter, edited.diameter)
    assert edited.metadata["adv_contact_angle_is_override"] is True


def test_xls_sidecar_exists_for_future_exact_alignment() -> None:
    assert (DATA_ROOT / "M5G2.XLS").exists()


def test_export_selected_results_to_one_xlsx(tmp_path: Path) -> None:
    results = [load_smp(DATA_ROOT / "M5G2.SMP"), load_smp(DATA_ROOT / "1-6.SMP")]
    output_path = export_results_xlsx(results, tmp_path / "selected.xlsx")

    assert output_path.exists()

    from openpyxl import load_workbook

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    assert workbook.sheetnames == ["Summary", "M5G2", "1-6"]
    summary = workbook["Summary"]
    assert summary["A5"].value == "File"
    assert summary["A6"].value == "M5G2"
    sample_sheet = workbook["M5G2"]
    headers = [sample_sheet.cell(14, column).value for column in range(1, 10)]
    assert "dV/dlogD (smoothed, intrusion only)" in headers
    assert all("raw" not in str(header).lower() for header in headers)
    workbook.close()


def test_subset626_sample_operator_submitter_metadata() -> None:
    smp = parse_smp(DATA_ROOT / "1号.SMP")

    assert smp.sample_name == "1号"
    assert smp.operator == "machenglong"
    assert smp.submitter == "machenglong"
